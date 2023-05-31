import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook

count = 2
urls = [
    'https://www.carrefour.com.ar/Electro-y-tecnologia',
    'https://www.carrefour.com.ar/Bazar-y-textil',
    'https://www.carrefour.com.ar/Almacen',
    'https://www.carrefour.com.ar/Desayuno-y-merienda',
    'https://www.carrefour.com.ar/Bebidas',
    'https://www.carrefour.com.ar/Lacteos-y-productos-frescos',
    'https://www.carrefour.com.ar/Carnes-y-Pescados',
    'https://www.carrefour.com.ar/Frutas-y-Verduras',
    'https://www.carrefour.com.ar/Panaderia',
    'https://www.carrefour.com.ar/Congelados',
    'https://www.carrefour.com.ar/Limpieza',
    'https://www.carrefour.com.ar/Perfumeria',
    'https://www.carrefour.com.ar/Mundo-Bebe',
    'https://www.carrefour.com.ar/Mascotas',
    'https://www.carrefour.com.ar/Bazar-y-textil/Indumentaria',
]

wb = Workbook()
ws = wb.active
ws['A1'] = 'Categoria'
ws['B1'] = 'Cantidad Productos'

options = Options()
options.add_argument('--headless')

driver = webdriver.Chrome(executable_path="chromedriver.exe", chrome_options=options)

for url in urls:
    driver.get(url)
    time.sleep(10)

    data = driver.find_element(By.CLASS_NAME, 'valtech-carrefourar-search-result-0-x-totalProducts--layout').text.split(' ')[0]
    label = url.split('/')[-1].replace('-', ' ')

    ws['A' + str(count)] = label
    ws['B' + str(count)] = data

    count += 1

wb.save("carrefour.xlsx")